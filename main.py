import pandas as pd
import os, datetime, calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

notation_dict = {'x': '8', 'v': 'Szabadság', 's': 'Betegség', 'p': 'Fizetett, igazolt távollét',
                 'u': 'Fizetés nélküli szabadság', 'r': 'Rendkívüli szabadnap', 'a': 'Apanap'}
days_dict = {0: "Hétfő", 1: "Kedd", 2: "Szerda", 3: "Csütörtök", 4: "Péntek", 5: "Szombat", 6: "Vasárnap"}

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

contents = os.listdir("input")
files = [os.path.join("input", content) for content in contents if os.path.isfile(os.path.join("input", content))]

for file in files:
    sheet_names = pd.ExcelFile(file).sheet_names
    for sheet in sheet_names:
        for char in sheet:
            if not char.isnumeric():
                delimiter = char
                break
        else:
            print(f"Couldn't find delimiter in sheet name {sheet} in file {file}.\n"
                  "Make sure sheet name resembles 'YYYY.mm'")
            continue

        year, month = [int(element) for element in sheet.split(delimiter)]
        number_of_days = calendar.monthrange(year=year, month=month)[1]
        dates = [datetime.date(year=year, month=month, day=day) for day in range(1, number_of_days + 1)]

        dating = input(f"{file}--->{sheet} sheet keltezése: ")

        df = pd.read_excel(file, sheet_name=sheet, header=2)
        df.drop(df.tail(9).index, inplace=True)
        columns_to_drop = [column for column in list(df.columns) if
                           "Unnamed" in str(column) or column not in ["Név", "Munkanap", "Ledolgozott napok",
                                                                      "Távollét", *[index for index in range(1, 32)]]]
        df.drop(columns=columns_to_drop, inplace=True)
        employees = df["Név"][1:]

        wb = Workbook()
        for employee in employees:
            index = df["Név"][df["Név"] == employee].index.values[0]
            employee_tuple = list(df.iterrows())[index]
            employee_data = employee_tuple[1].values
            employee_dict = dict(zip(dates, employee_data[1:-3]))
            ws = wb.create_sheet()

            ws.merge_cells("A1:D1")
            ws["A1"] = "Munkaidő nyilvántartás"
            ws["A1"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["A1"].font = Font(name='Calibri', size=16, bold=True)

            ws.merge_cells("A2:D3")
            ws["A2"] = "Glownexus Hungary Kft."
            ws["A2"].fill = PatternFill(fill_type="solid", start_color="b0d4e4", end_color="b0d4e4")
            ws["A2"].font = Font(name='Calibri', size=14, bold=True)

            ws["B4"] = "Név"
            ws["B4"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["B4"].font = Font(name='Calibri', size=11, bold=True)

            ws.merge_cells("C4:D4")
            ws["C4"] = employee
            ws["C4"].fill = PatternFill(fill_type="solid", start_color="d0ecfc", end_color="d0ecfc")
            ws["C4"].font = Font(name='Verdana', size=14, bold=True, color="c72918")

            ws["A4"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")

            ws.merge_cells("A5:A6")
            ws["A5"] = "Dátum"
            ws["A5"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["A5"].font = Font(name='Calibri', size=11, bold=True)

            ws.merge_cells("B5:B6")
            ws["B5"] = "Nap"
            ws["B5"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["B5"].font = Font(name='Calibri', size=11, bold=True)

            ws.merge_cells("C5:C6")
            ws["C5"] = "Ledolgozott órák"
            ws["C5"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["C5"].font = Font(name='Calibri', size=11, bold=True)

            ws["D5"] = "Szabadság"
            ws["D5"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["D5"].font = Font(name='Calibri', size=11, bold=True)

            ws["D6"] = "Betegség"
            ws["D6"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
            ws["D6"].font = Font(name='Calibri', size=11, bold=True)

            for i, key in enumerate(employee_dict):
                value = employee_dict[key]
                ws[f"A{7 + i}"] = key.strftime("%Y/%m/%d")
                ws[f"A{7 + i}"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
                ws[f"B{7 + i}"] = days_dict[key.weekday()]
                ws[f"B{7 + i}"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
                ws[f"C{7 + i}"].alignment = Alignment(horizontal='center')
                ws[f"D{7 + i}"].alignment = Alignment(horizontal='center')
                if value == "x":
                    ws[f"C{7 + i}"] = "8"
                elif isinstance(value, float):
                    ws[f"C{7 + i}"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
                    ws[f"D{7 + i}"].fill = PatternFill(fill_type="solid", start_color="b0acac", end_color="b0acac")
                else:
                    ws[f"D{7 + i}"] = notation_dict[value]

            ws[f"A{7 + i + 1}"] = "Munkanap összesen:"
            ws[f"C{7 + i + 1}"] = employee_data[-2]
            ws[f"C{7 + i + 1}"].alignment = Alignment(horizontal='center')
            ws[f"D{7 + i + 1}"] = employee_data[-1]
            ws[f"D{7 + i + 1}"].alignment = Alignment(horizontal='center')

            ws.merge_cells(f"C{7 + i + 3}:D{7 + i + 3}")
            ws[f"C{7 + i + 3}"] = "A fenti kimutatást elfogadom"
            ws[f"C{7 + i + 3}"].alignment = Alignment(horizontal='center')

            ws.merge_cells(f"A{7 + i + 5}:B{7 + i + 5}")
            ws[f"A{7 + i + 5}"] = f"Budapest, {dating}"
            ws[f"A{7 + i + 5}"].alignment = Alignment(horizontal='center')

            ws.merge_cells(f"C{7 + i + 5}:D{7 + i + 5}")
            ws[f"C{7 + i + 5}"] = "_____________________________"
            ws[f"C{7 + i + 5}"].alignment = Alignment(horizontal='center')

            ws.merge_cells(f"C{7 + i + 6}:D{7 + i + 6}")
            ws[f"C{7 + i + 6}"] = employee
            ws[f"C{7 + i + 6}"].font = Font(name='Tahoma', size=12, bold=True, color="c72918")
            ws[f"C{7 + i + 6}"].alignment = Alignment(horizontal='center')

            for column in ["A", "B", "C", "D"]:
                ws.column_dimensions[column].width = 20
                for row in range(1, 7 + i + 1):
                    ws[f"{column}{row}"].border = thin_border
                for row in range(1, 7):
                    ws[f"{column}{row}"].alignment = Alignment(horizontal='center')

        del wb[wb.sheetnames[0]]
        wb.save(os.path.join("output", f"{year}.{month}.xlsx"))
